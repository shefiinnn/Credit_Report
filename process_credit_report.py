#!/usr/bin/env python
import logging
import os
import re
import json
import sys
from datetime import datetime
import PyPDF2
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Set pdfplumber's logger to ERROR level to reduce noise
logging.getLogger("pdfplumber").setLevel(logging.ERROR)

class CreditReportProcessor:
    def __init__(self, pdf_path=None, output_dir="output"):
        """Initialize the CreditReportProcessor with default parameters"""
        self.pdf_path = pdf_path
        self.output_dir = output_dir
        self.json_output = os.path.join(output_dir, "credit_report.json")
        self.excel_output = os.path.join(output_dir, "credit_report.xlsx")
        self.initialize_credit_report()
        
    def initialize_credit_report(self):
        """Initialize the credit report data structure"""
        self.credit_report = {
            "personal_info": {"transunion": {}, "experian": {}, "equifax": {}},
            "summary": {"transunion": {}, "experian": {}, "equifax": {}},
            "accounts": [],
            "collections": [],  # Added specific section for collections
            "inquiries": [],
            "public_information": {"transunion": {},"experian": {},"equifax": {}},

        }

    def process(self, pdf_path=None, output_dir=None):
        """Process the credit report PDF from start to finish"""
        if pdf_path:
            self.pdf_path = pdf_path
        if output_dir:
            self.output_dir = output_dir
            self.json_output = os.path.join(output_dir, "credit_report.json")
            self.excel_output = os.path.join(output_dir, "credit_report.xlsx")

        if not self.pdf_path:
            raise ValueError("PDF path is required.")

        os.makedirs(self.output_dir, exist_ok=True)
        self.initialize_credit_report()

        print(f"Processing credit report: {self.pdf_path}")

        # ✅ Extract both text and page
        text, page = self.extract_lines_from_pdf()
        if not text:
            print("Failed to extract text from PDF")
            return False

        # ✅ Pass both to parser
        print("Parsing credit report data...")
        self.parse_credit_report(text, page)

        # Save data to JSON
        print(f"Saving JSON to {self.json_output}...")
        self.save_json()

        # Create Excel file with multiple sheets
        print(f"Creating Excel file {self.excel_output}")
        self.create_excel()

        print("Processing complete!")
        
        return True
    
    def extract_lines_from_pdf(self):
        """Extract text from PDF using pdfplumber and return both text and first page object"""
        if not os.path.exists(self.pdf_path):
            print(f"Error: PDF file {self.pdf_path} not found")
            return None, None  # Return both values

        print("Extracting text from PDF...")

        try:
            with pdfplumber.open(self.pdf_path) as pdf:
                text = ""
                first_page = pdf.pages[0] if len(pdf.pages) > 0 else None

                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n\n"

            if len(text) > 1000 and first_page:
                print("Successfully extracted text with pdfplumber")
                return text, first_page  # ✅ return both
        except Exception as e:
            print(f"pdfplumber extraction error: {e}")

        # Fall back to PyPDF2 (though coordinates will not be supported)
        try:
            with open(self.pdf_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                text = ""
                for page in reader.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n\n"

            if len(text) > 1000:
                print("Successfully extracted text with PyPDF2")
                return text, None  # No page object available
            else:
                print("Text extraction yielded insufficient content")
        except Exception as e:
            print(f"PyPDF2 extraction error: {e}")

        return None, None 

    def parse_credit_report(self, text,page):
        """Parse the entire credit report text into structured data"""
        lines = text.splitlines()
        self.extract_personal_info(lines,page)
        self.extract_summary(text)
        self.extract_accounts_data(text)
        self.extract_collections(text)  
        self.extract_inquiries(text)
        self.extract_public_information(text)
        self.update_inquiry_counts()
    
    def extract_personal_info(self, lines, page):
        """Extract personal information from lines and pdfplumber page"""
        import re
        import json
        from collections import defaultdict
        from sklearn.cluster import KMeans
        import numpy as np

        bureaus = ["transunion", "experian", "equifax"]
        words = page.extract_words()
        self.credit_report["personal_info"] = {
            "transunion": {},
            "experian": {},
            "equifax": {}
        }

        # ✅ Extract Name
        name_start = name_end = None

        for word in words:
            text = word["text"].lower()
            if "name" in text and name_start is None:
                name_start = word["top"] - 1
            elif name_start is not None and any(k in text for k in ["also", "date", "consumer", "transunion", "equifax", "experian", "https", "statement"]):
                name_end = word["top"]
                break

        if name_start is None or name_end is None:
            print("⚠️ Could not locate name section.")
        else:
            name_words = [
                w for w in words
                if name_start <= w["top"] < name_end and w["text"].lower() != "name"
            ]

            if len(name_words) < 6:
                print("⚠️ Not enough name words to cluster.")
            else:
                # Cluster by x0 to map to bureaus
                name_x = np.array([[w["x0"]] for w in name_words])
                kmeans = KMeans(n_clusters=3, random_state=42).fit(name_x)
                labels = kmeans.labels_

                # Map clusters to bureaus
                cluster_centers = sorted([(i, c[0]) for i, c in enumerate(kmeans.cluster_centers_)], key=lambda x: x[1])
                cluster_to_bureau = {
                    cluster_centers[0][0]: "transunion",
                    cluster_centers[1][0]: "experian",
                    cluster_centers[2][0]: "equifax"
                }

                # Group vertically like address/employer
                name_lines = {
                    "transunion": defaultdict(list),
                    "experian": defaultdict(list),
                    "equifax": defaultdict(list)
                }

                def align_name_top(top, line_dict, tolerance=2.5):
                    for existing in line_dict:
                        if abs(existing - top) <= tolerance:
                            return existing
                    return top

                for i, word in enumerate(name_words):
                    bureau = cluster_to_bureau[labels[i]]
                    top = align_name_top(word["top"], name_lines[bureau])
                    name_lines[bureau][top].append((word["x0"], word["text"]))

                for bureau in bureaus:
                    bureau_lines = name_lines[bureau]
                    sorted_lines = sorted(bureau_lines.items())
                    all_words = []

                    for _, word_group in sorted_lines:
                        sorted_words = [w for _, w in sorted(word_group)]
                        all_words.append(" ".join(sorted_words))

                    final_name = "  ".join(all_words).strip()
                    if not final_name:
                        final_name = "N/A"

                    self.credit_report["personal_info"][bureau]["name"] = final_name
                    print(f"✅ Extracted name for {bureau}: {final_name}")

        # ✅ Extract DOB
            for line in lines:
                if isinstance(line, str):  # ✅ Make sure it's a string
                    dob_match = re.search(r'Date of Birth\s+(\d{4})', line)
                    if dob_match:
                        for bureau in bureaus:
                            self.credit_report["personal_info"][bureau]["year_of_birth"] = dob_match.group(1)
                        print(f"✅ Found birth year: {dob_match.group(1)}")
                        break

        # ✅ Extract Current Address (TransUnion)
        for i, line in enumerate(lines):
            if "Current Address" in line:
                street_line = line.replace("Current Address", "").strip()
                city_line = lines[i + 1].strip() if i + 1 < len(lines) else ""
                street_parts = street_line.split()
                city_parts = city_line.split()
                if len(street_parts) >= 3 and len(city_parts) >= 3:
                    s_chunk = len(street_parts) // 3
                    c_chunk = len(city_parts) // 3
                    t_street = " ".join(street_parts[:s_chunk])
                    t_city = " ".join(city_parts[:c_chunk])
                    self.credit_report["personal_info"]["transunion"]["current_address"] = f"{t_street} {t_city}".strip()
                    print("✅ Final TransUnion address:", self.credit_report["personal_info"]["transunion"]["current_address"])
                else:
                    print("⚠️ Address format unexpected; skipping.")
                break

        # ✅ Extract Report Date
        for line in lines:
            report_date_match = re.search(r'Credit Report Date\s+(\d{1,2}/\d{1,2}/\d{4})', line)
            if report_date_match:
                for bureau in bureaus:
                    self.credit_report["personal_info"][bureau]["report_date"] = report_date_match.group(1)
                print(f"✅ Found Credit Report Date: {report_date_match.group(1)}")

        # ✅ Extract Previous Addresses (KMeans logic)
        prev_start = prev_end = None

        SECTION_TRIGGERS = {"employer", "date", "consumer", "transunion", "experian", "equifax", "statement", "https"}

        for word in words:
            text = word["text"].lower()
            if any(key in text for key in ["previous", "prior"]) and prev_start is None:
                prev_start = word["top"] - 1
            elif prev_start is not None and any(trigger in text for trigger in SECTION_TRIGGERS):
                prev_end = word["top"] - 2  # end right before the new section
                break

        addr_words = [
            w for w in words
            if prev_start < w["top"] < prev_end
            and w["text"].lower() not in {"previous", "address"}
]

        if len(addr_words) < 9:
            print("⚠️ Not enough words found to apply clustering.")
            return

        x_values = np.array([[w["x0"]] for w in addr_words])
        kmeans = KMeans(n_clusters=3, random_state=42).fit(x_values)
        labels = kmeans.labels_

        # Map cluster index to bureau name based on x-position
        cluster_centers = sorted([(i, c[0]) for i, c in enumerate(kmeans.cluster_centers_)], key=lambda x: x[1])
        cluster_to_bureau = {
            cluster_centers[0][0]: "transunion",
            cluster_centers[1][0]: "experian",
            cluster_centers[2][0]: "equifax"
        }

        grouped = {
            "transunion": defaultdict(list),
            "experian": defaultdict(list),
            "equifax": defaultdict(list)
        }

        line_tolerance = 2.5

        def get_aligned_top(top, line_dict):
            for existing_top in line_dict:
                if abs(existing_top - top) <= line_tolerance:
                    return existing_top
            return top

        for i, word in enumerate(addr_words):
            bureau = cluster_to_bureau[labels[i]]
            aligned_top = get_aligned_top(word["top"], grouped[bureau])
            grouped[bureau][aligned_top].append((word["x0"], word["text"]))

        for bureau in bureaus:
            lines = grouped[bureau]
            sorted_lines = sorted(lines.items())
            address_lines = []
            for _, word_group in sorted_lines:
                sorted_words = [w for _, w in sorted(word_group)]
                address_lines.append(" ".join(sorted_words))
            full_address = "  ".join(address_lines)
            self.credit_report["personal_info"][bureau]["previous_address"] = full_address
            print(f"✅ Previous address for {bureau}: {full_address}")
        # ✅ Employer block start
        emp_start = emp_end = None

        for word in words:
            text = word["text"].lower()
            if "employer" in text and emp_start is None:
                emp_start = word["top"] - 1
            elif "consumer" in text and emp_start is not None:
                emp_end = word["top"]
                break

        if emp_start is None or emp_end is None:
            print("⚠️ Could not find Employer section boundaries.")
            return

        emp_words = [
            w for w in words
            if emp_start <= w["top"] < emp_end and w["text"].lower() != "employer"
        ]

        if len(emp_words) < 9:
            print("⚠️ Not enough words to apply clustering to employer section.")
            return

        emp_x = np.array([[w["x0"]] for w in emp_words])
        kmeans = KMeans(n_clusters=3, random_state=42).fit(emp_x)
        labels = kmeans.labels_

        # Map clusters to bureaus (left to right)
        cluster_centers = sorted([(i, center[0]) for i, center in enumerate(kmeans.cluster_centers_)], key=lambda x: x[1])
        cluster_to_bureau = {
            cluster_centers[0][0]: "transunion",
            cluster_centers[1][0]: "experian",
            cluster_centers[2][0]: "equifax"
        }

        # Vertical clustering just like previous address
        employer_lines = {
            "transunion": defaultdict(list),
            "experian": defaultdict(list),
            "equifax": defaultdict(list)
        }

        def get_aligned_top(top, line_dict, tolerance=2.5):
            for existing in line_dict:
                if abs(existing - top) <= tolerance:
                    return existing
            return top

        for i, word in enumerate(emp_words):
            x0 = word["x0"]
            text=word["text"]
            if "service" in text.lower():
                print(f"SERVICES → x0: {x0} top: {word['top']}")

            # Optional override if needed
            if x0 < 280:
                bureau = "transunion"
            else:
                label=labels[i]
                bureau = cluster_to_bureau[label]

            aligned_top = get_aligned_top(word["top"], employer_lines[bureau])
            employer_lines[bureau][aligned_top].append((x0, word["text"]))

        # Final parsing logic
        def is_date_line(text):
            return re.fullmatch(r"\d{1,2}/\d{1,2}/\d{2,4}", text.strip()) is not None

        def is_date_updated_tag(text):
            return "date updated" in text.lower()

        def is_valid_employer_name(text):
            return bool(re.search(r"[A-Za-z]{2,}", text))  # at least 2 alphabetic characters

        for bureau in bureaus:
            lines = employer_lines[bureau]
            sorted_lines = sorted(lines.items())
            full_lines = []

            last_index = -1
            last_employer_line = ""

            for _, word_group in sorted_lines:
                line = " ".join(w for _, w in sorted(word_group)).strip()
                if not line:
                    continue

                # Handle date update lines: extract date only
                if is_date_updated_tag(line) and last_index != -1:
                    match = re.search(r"\d{1,2}/\d{1,2}/\d{2,4}", line)
                    if match:
                        full_lines[last_index] += f"  {match.group()}"
                    continue

                if is_date_line(line):
                    continue  # Skip lone dates

                if is_valid_employer_name(line):
                    if line != last_employer_line:
                        full_lines.append(line)
                        last_index = len(full_lines) - 1
                        last_employer_line = line
                    continue

            full_employer_block = "  ".join(full_lines).strip()
            if not full_employer_block or re.fullmatch(r"[-–/\s]*", full_employer_block):
                full_employer_block = "N/A"

            self.credit_report["personal_info"][bureau]["employer"] = full_employer_block
            print(f"✅ Final employer block for {bureau}: {full_employer_block}")
    def extract_summary(self, text):
        """Extract summary information from the text"""
        # Dictionary mapping patterns to keys
        summary_patterns = {
            r'Total Accounts\s+(\d+)\s+(\d+)\s+(\d+)': "total_accounts",
            r'Open Accounts:\s+(\d+)\s+(\d+)\s+(\d+)': "open_accounts",
            r'Closed Accounts:\s+(\d+)\s+(\d+)\s+(\d+)': "closed_accounts",
            r'Delinquent:\s+(\d+)\s+(\d+)\s+(\d+)': "delinquent",
            r'Derogatory:\s+(\d+)\s+(\d+)\s+(\d+)': "derogatory",
            r'Balances:\s+\$([0-9,]+)\s+\$([0-9,]+)\s+\$([0-9,]+)': "balances",
            r'(?:Monthly )?Payments?:\s+\$([0-9,]+)\s+\$([0-9,]+)\s+\$([0-9,]+)': "monthly_payments",
            r'Credit Utilization:\s+(\d+)%\s+(\d+)%\s+(\d+)%': "credit_utilization",
            r'Public Records:\s+(\d+)\s+(\d+)\s+(\d+)': "public_records",
            r'Inquiries\s+\(2\s+years?\):\s+(\d+)\s+(\d+)\s+(\d+)': "inquiries_2y",
            r'Inquiries\s+\(\d+\s+years?\):\s+(\d+)\s+(\d+)\s+(\d+)': "inquiries_alt"
        }
        
        bureaus = ["transunion", "experian", "equifax"]
        
        for pattern, key in summary_patterns.items():
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                for i, bureau in enumerate(bureaus):
                    if key == "inquiries_alt" and "inquiries_2y" not in self.credit_report["summary"][bureau]:
                        self.credit_report["summary"][bureau]["inquiries_2y"] = match.group(i+1)
                    else:
                        self.credit_report["summary"][bureau][key] = match.group(i+1)
                    
        # Try alternative inquiry pattern if the first one doesn't match
        if not any("inquiries_2y" in self.credit_report["summary"][bureau] for bureau in bureaus):
            alt_inquiry_patterns = [
                r'Inquiries\s+\(\d+ yr\):\s+(\d+)\s+(\d+)\s+(\d+)',
                r'Inquiries:\s+(\d+)\s+(\d+)\s+(\d+)'
            ]
            
            for pattern in alt_inquiry_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    for i, bureau in enumerate(bureaus):
                        self.credit_report["summary"][bureau]["inquiries_2y"] = match.group(i+1)
                    break

    def update_inquiry_counts(self):
        """Update summary with inquiry counts if they weren't found in the extract_summary phase"""
        bureaus = ["transunion", "experian", "equifax"]
        
        # Check if we already have inquiry counts in the summary
        if any("inquiries_2y" in self.credit_report["summary"][bureau] for bureau in bureaus):
            return
            
        # Count inquiries by bureau
        counts = {"transunion": 0, "experian": 0, "equifax": 0, "unknown": 0}
        
        for inquiry in self.credit_report["inquiries"]:
            bureau = inquiry.get("bureau", "unknown").lower()
            if bureau in counts:
                counts[bureau] += 1
            else:
                counts["unknown"] += 1
        
        # If we have an "unknown" category but no others, distribute evenly
        if counts["unknown"] > 0 and all(counts[b] == 0 for b in bureaus):
            # Look for summary numbers in the provided data
            if "inquiries_count" in self.credit_report.get("summary_data", {}):
                for i, bureau in enumerate(bureaus):
                    count = self.credit_report["summary_data"]["inquiries_count"][i]
                    self.credit_report["summary"][bureau]["inquiries_2y"] = str(count)
            else:
                # No counts available, use the total number for all bureaus as a fallback
                for bureau in bureaus:
                    self.credit_report["summary"][bureau]["inquiries_2y"] = str(len(self.credit_report["inquiries"]))
        else:
            # Use the counts we calculated
            for bureau in bureaus:
                self.credit_report["summary"][bureau]["inquiries_2y"] = str(counts[bureau])

    def extract_inquiries(self, text):
        """Extract inquiries data from text with improved pattern matching"""
        print("Extracting inquiries...")
        
        # First, try to extract summary inquiry counts directly from the text
        inquiry_count_pattern = r'Inquiries\s+\(2\s+years?\):\s+(\d+)\s+(\d+)\s+(\d+)'
        match = re.search(inquiry_count_pattern, text, re.IGNORECASE)
        if match:
            # Store these counts to use later if needed
            self.credit_report["summary_data"] = {
                "inquiries_count": [int(match.group(1)), int(match.group(2)), int(match.group(3))]
            }
            print(f"Found inquiry counts: TU={match.group(1)}, EX={match.group(2)}, EQ={match.group(3)}")
        
        # Find inquiries section
        inquiries_section = None
        inquiries_header_patterns = [
            r'Inquiries\s*(?:Creditor Name|Date|Credit Bureau).*',
            r'(?:Hard|Soft)?\s*Inquiries\s*'
        ]
        
        for pattern in inquiries_header_patterns:
            inquiry_match = re.search(pattern, text, re.IGNORECASE)
            if inquiry_match:
                start_idx = inquiry_match.end()
                # Look for a reasonable end point (next major section or end of document)
                end_markers = [
                    r'\n\s*(?:Accounts|Public Records|Credit Score|Summary|Report Details)',
                    r'\n\s*Page \d+ of \d+',
                    r'\n\s*Creditor Contacts',
                    r'\n\s*© myfreescorenow\.com'
                ]
                
                end_idx = len(text)
                for end_pattern in end_markers:
                    end_match = re.search(end_pattern, text[start_idx:], re.IGNORECASE)
                    if end_match:
                        end_idx = start_idx + end_match.start()
                        break
                
                inquiries_section = text[start_idx:end_idx]
                break
        
        if not inquiries_section:
            # Try to find the Inquiries section by looking for "Creditor Name Date of Inquiry Credit Bureau"
            section_match = re.search(r'Creditor Name\s+Date of Inquiry\s+Credit Bureau\s*(.*?)(?:(?:Creditor Contacts)|©)', text, re.DOTALL)
            if section_match:
                inquiries_section = section_match.group(1)
                print(f"Found inquiries section using alternative method, {len(inquiries_section)} characters")
            else:
                print("No inquiries section found")
                return
        else:
            print(f"Found inquiries section of {len(inquiries_section)} characters")
        
        # Extract inquiries in tabular format
        lines = inquiries_section.split('\n')
        for line in lines:
            line = line.strip()
            if not line or len(line) < 5 or re.match(r'^Page \d+|https?://', line):
                continue
                
            # Try to match the inquiry format seen in the PDF
            inquiry_match = re.match(r'([A-Z0-9\s\/#\-\.\,&]+)\s+(\d{1,2}/\d{1,2}/\d{4})\s+(TransUnion|Experian|Equifax)',line
)

            if inquiry_match:
                creditor = inquiry_match.group(1).strip()
                date = self.clean_value(inquiry_match.group(2))
                bureau = inquiry_match.group(3).lower()
                
                # Skip if this looks like a header row
                if creditor.lower() in ['creditor name', 'inquiry', 'inquiries']:
                    continue
                
                self.credit_report["inquiries"].append({
                    "creditor": creditor,
                    "date": date,
                    "bureau": bureau
                })
        
        print(f"Extracted {len(self.credit_report['inquiries'])} inquiries")
    
    def extract_collections(self, text):
        """Extract collection accounts from the text"""
        print("Extracting collection accounts...")
        collection_sections = []
        
        # Find the Collection section headers
        collection_pattern = r'Collection\s*\n([A-Z][A-Z0-9\s&.,\'"-/]+?)\s*\n'
        collection_matches = re.finditer(collection_pattern, text)
        
        for match in collection_matches:
            start_idx = match.start()
            next_section_match = re.search(r'\n([A-Z][A-Z0-9\s&.,\'"-/]+?)\s*\n', text[start_idx + 10:])
            
            if next_section_match:
                end_idx = start_idx + 10 + next_section_match.start()
            else:
                end_idx = start_idx + 1000  # Arbitrary limit if no next section found
                
            collection_text = text[start_idx:end_idx]
            collection_sections.append(collection_text)
        
        # If no collection sections found, try to find collection accounts in the account listings
        if not collection_sections:
            # Check for collection agencies in the accounts data
            collection_agencies = [
                "NATLCRSYS", "IQ DATA INT", "CR SYST INTR", "RADIUSGLOBAL", "ALDOUS"
            ]
            
            for agency in collection_agencies:
                agency_match = re.search(fr'{agency}\s*\n(.*?)\n(?:[A-Z][A-Z0-9\s&.,\'"-/]+?|\Z)', text, re.DOTALL)
                if agency_match:
                    collection_sections.append(f"Collection\n{agency}\n{agency_match.group(1)}")
        
        # Process each collection section
        for collection_text in collection_sections:
            try:
                collection_data = self.parse_collection(collection_text)
                if collection_data:
                    self.credit_report["collections"].append(collection_data)
            except Exception as e:
                print(f"Error parsing collection: {e}")
                
        print(f"Extracted {len(self.credit_report['collections'])} collection accounts")
    
    def parse_collection(self, collection_text):
        """Parse a collection account section into structured data"""
        # Extract the collection agency name
        agency_match = re.search(r'Collection\s*\n([A-Z][A-Z0-9\s&.,\'"-/]+?)\s*\n', collection_text)
        if not agency_match:
            agency_match = re.search(r'^([A-Z][A-Z0-9\s&.,\'"-/]+?)\s*\n', collection_text)
            
        if not agency_match:
            return None
            
        agency = agency_match.group(1).strip()
        
        # Initialize collection data
        collection_data = {
            "agency": agency,
            "transunion": {},
            "experian": {},
            "equifax": {}
        }
        
        # Define key-value patterns to search for
        key_value_patterns = [
            (r'Account #\s*([^\n]*)', "account_number"),
            (r'High Balance:\s*\$?([0-9,.]+)', "high_balance"),
            (r'Balance Owed:\s*\$?([0-9,.]+)', "balance_owed"),
            (r'Date Reported:\s*([0-9/]+)', "date_reported"),
            (r'Date Opened:\s*([0-9/]+)', "date_opened"),
            (r'Original Creditor\s*([^\n]*)', "original_creditor"),
            (r'Last Payment:\s*([0-9/]*)', "last_payment")
        ]
        
        # Extract data for each key pattern
        for pattern, key in key_value_patterns:
            matches = re.findall(pattern, collection_text)
            if matches:
                bureaus = ["transunion", "experian", "equifax"]
                
                if len(matches) >= 3:
                    # If we have three values, assign one to each bureau
                    for i, bureau in enumerate(bureaus):
                        if i < len(matches):
                            collection_data[bureau][key] = self.clean_value(matches[i])
                elif len(matches) == 1:
                    # If only one value found, assign it to all bureaus
                    value = self.clean_value(matches[0])
                    for bureau in bureaus:
                        collection_data[bureau][key] = value
        
        # Check if we found any useful data
        has_data = False
        for bureau in ["transunion", "experian", "equifax"]:
            if collection_data[bureau]:
                has_data = True
                break
                
        return collection_data if has_data else None
    
    def extract_accounts_data(self, text):
        """Extract all account data from the text using improved pattern matching"""
        print("Extracting account data...")
        
        # Split by lines for processing
        lines = text.split('\n')
        
        # Find potential creditor sections
        account_sections = self.identify_account_sections(lines)
        
        # Parse each account section
        for account_text in account_sections:
            try:
                account_data = self.parse_account(account_text)
                if account_data:
                    self.credit_report["accounts"].append(account_data)
            except Exception as e:
                print(f"Error parsing account: {e}")
        
        print(f"Extracted {len(self.credit_report['accounts'])} accounts")
        
    def identify_account_sections(self, lines):
        """Identify sections that contain account information"""
        creditor_indices = []
        account_sections = []
        
        # Patterns to ignore when finding creditor names
        ignore_patterns = [
            r'transunion®\s+experian®\s+equifax®',
            r'account\s+#',
            r'high\s+balance',
            r'page\s+\d+\s+of\s+\d+',
            r'collection',
            r'past\s+due\s+amount',
            r'two-year\s+payment\s+history',
            r'inquiries',
            r'https?://',
            r'days late',
            r'creditor contacts'
        ]
        ignore_regex = re.compile('|'.join(ignore_patterns), re.IGNORECASE)
        
        i = 0
        while i < len(lines) - 5:
            line = lines[i].strip()
            
            # Skip empty or ignored lines
            if not line or ignore_regex.search(line):
                i += 1
                continue
            
            # Check if line is a potential creditor name (all caps, no digits, reasonable length)
            if (re.match(r'^[A-Z][A-Z0-9\s&.,\'"-/]+$', line) and  
                len(line) < 40 and
                not any(char.isdigit() for char in line) and
                not line.startswith("Page") and
                not line.startswith("http")):
                
                # Look ahead to see if this is followed by account information
                next_text = ' '.join(lines[i+1:i+10]).lower()
                
                # Check for account information markers
                if (('account #' in next_text or 'high balance' in next_text) and 
                    any(term in next_text for term in ['balance', 'opened', 'status', 'payment'])):
                    
                    creditor_indices.append((i, line))
                
            i += 1
        
        # Extract account sections based on creditor positions
        for idx, (start_idx, creditor) in enumerate(creditor_indices):
            # Find end of section (next creditor or end marker)
            if idx < len(creditor_indices) - 1:
                end_idx = creditor_indices[idx+1][0]
            else:
                # For the last account, look for end markers
                end_idx = len(lines)
                for j in range(start_idx + 10, min(start_idx + 100, len(lines))):
                    if j >= len(lines):
                        break
                    if (re.match(r'^Page \d+ of \d+', lines[j].strip()) or
                        re.match(r'^https?://', lines[j].strip()) or
                        re.match(r'^[A-Z][A-Z\s]+(?:INQUIRIES|SUMMARY|REPORT|COLLECTION)', lines[j].strip()) or
                        re.match(r'^Creditor Contacts', lines[j].strip())):
                        end_idx = j
                        break
                
            # Extract the account section
            section = '\n'.join(lines[start_idx:end_idx])
            print(f"Found account: {creditor} (lines {start_idx+1} to {end_idx})")
            account_sections.append(section)
        
        return account_sections
    
    
    def parse_account(self, account_text):
        lines = account_text.split('\n')
        creditor_lines = []
        for i, line in enumerate(lines):
            line = line.strip()
            if not line:
                continue
            if i == 0 or (re.match(r'^[A-Z\s&.,\'"-]+$', line) and not re.search(r'Account #|High Balance|Balance Owed', line, re.IGNORECASE)):
                creditor_lines.append(line)
            else:
                break
        
        creditor = ' '.join(creditor_lines).strip()
        
        account_data = {
            "creditor": creditor,
            "transunion": {},
            "experian": {},
            "equifax": {},
            "payment_history": { 
                "transunion": {"months": [], "statuses": []}, 
                "experian": {"months": [], "statuses": []}, 
                "equifax": {"months": [], "statuses": []} 
            }
        }
        
        key_value_patterns = [
            (r'Account #\s*([^\n]*)', "account_number"),
            (r'High Balance:\s*([^\n]*)', "high_balance"),
            (r'Balance Owed:\s*([^\n]*)', "balance_owed"),
            (r'Account Status:\s*([^\n]*)', "account_status"),
            (r'Payment Status:\s*([^\n]*)', "payment_status"),
            (r'Date Opened:\s*([^\n]*)', "date_opened"),
            (r'Last Payment:\s*([^\n]*)', "last_payment"),
            (r'Account Type:\s*([^\n]*)', "account_type"),
            (r'Credit Limit:\s*([^\n]*)', "credit_limit"),
            (r'Term Length:\s*([^\n]*)', "term_length"),
            (r'Creditor Type:\s*([^\n]*)', "creditor_type"),
            (r'Creditor Remarks:\s*([^\n]*)', "creditor_remarks"),
            (r'Date Reported:\s*([^\n]*)', "date_reported"),
            (r'Date of Last Activity:\s*([^\n]*)', "date_of_last_activity"),
            (r'Past Due Amount:\s*([^\n]*)', "past_due_amount"),
            (r'Last Verified:\s*([^\n]*)', "last_verified"),
            (r'Closed Date:\s*([^\n]*)', "closed_date"),
            (r'Account Rating:\s*([^\n]*)', "account_rating"),
            (r'Account Description:\s*([^\n]*)', "account_description"),
            (r'Dispute Status:\s*([^\n]*)', "dispute_status"),
            (r'Payment Amount:\s*([^\n]*)', "payment_amount"),
            (r'Payment Frequency:\s*([^\n]*)', "payment_frequency")
        ]
        
        for pattern, key in key_value_patterns:
            matches = re.search(pattern, account_text)
            if matches:
                value_text = matches.group(1).strip()
                values = value_text.split()
                
                bureaus = ["transunion", "experian", "equifax"]
                
                if len(values) >= 3:
                    for i, bureau in enumerate(bureaus):
                        if i < len(values):
                            account_data[bureau][key] = self.clean_value(values[i].strip())
                elif len(values) == 1:
                    for bureau in bureaus:
                        account_data[bureau][key] = self.clean_value(values[0].strip())
        
        # --- REFINED TWO-YEAR PAYMENT HISTORY EXTRACTION ---
        try:
            lines = account_text.splitlines()
            history_start_index = -1

            for i, line in enumerate(lines):
                if "Two-Year Payment History" in line:
                    history_start_index = i
                    break
            
            if history_start_index != -1:
                # Store all month and status tokens found within the history section for each bureau
                temp_extracted_months_per_bureau = {"transunion": [], "experian": [], "equifax": []}
                temp_extracted_statuses_per_bureau = {"transunion": [], "experian": [], "equifax": []}
                
                current_bureau = None
                
                # Month/Year pattern: Captures 'Jan', 'Feb', ..., 'Dec', and '\'YY' (like '24, '25)
                month_year_token_pattern = re.compile(r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec|\'\d{2})", re.IGNORECASE)
                
                # Status pattern: Prioritize 'NO DATA'. Explicitly exclude 'NO' alone.
                # If 'NO DATA' is not found, and only 'NO' appears, it implies a parsing error or missing data.
                # We want to force it to be 'NO DATA' if it was meant to be, or rely on padding.
                status_token_pattern = re.compile(r"(OK|30|60|90|120|150|CO|COLLECTION|NO\s*DATA)", re.IGNORECASE)

                # Iterate through lines *after* the "Two-Year Payment History" header
                for i in range(history_start_index + 1, min(len(lines), history_start_index + 100)):
                    line = lines[i].strip()
                    lower_line = line.lower()

                    # Stop parsing if we encounter a new section, ensuring no month/status pattern is found on this line
                    if (re.search(r'(days late|account description|summary|inquiries|public information|account terms|tradelines|previous addresses|credit score)', lower_line, re.IGNORECASE) and
                        not (month_year_token_pattern.search(line) or status_token_pattern.search(line))):
                        break

                    # Detect bureau header
                    if "transunion" in lower_line and "score" not in lower_line:
                        current_bureau = "transunion"
                        continue
                    elif "experian" in lower_line and "score" not in lower_line:
                        current_bureau = "experian"
                        continue
                    elif "equifax" in lower_line and "score" not in lower_line:
                        current_bureau = "equifax"
                        continue
                    
                    if current_bureau:
                        # Handle "NONE REPORTED" (e.g., if a bureau has no history at all)
                        if "none reported" in lower_line or "no history" in lower_line:
                            temp_extracted_months_per_bureau[current_bureau] = []
                            temp_extracted_statuses_per_bureau[current_bureau] = ["NONE REPORTED"]
                            current_bureau = None
                            continue

                        # Extract month/year tokens
                        found_months = month_year_token_pattern.findall(line)
                        temp_extracted_months_per_bureau[current_bureau].extend([m.upper() for m in found_months])

                        # Extract status tokens
                        found_statuses = status_token_pattern.findall(line)
                        # Process found statuses: If "NO" is found without "DATA", treat it as "NO DATA"
                        processed_statuses = []
                        for s in found_statuses:
                            if s.upper() == 'NO':
                                # This handles cases where OCR might split "NO DATA" or just see "NO"
                                # and we want it to be "NO DATA" for consistency.
                                processed_statuses.append('NO DATA')
                            else:
                                processed_statuses.append(s.upper())
                        temp_extracted_statuses_per_bureau[current_bureau].extend(processed_statuses)
                
                # --- Final Processing and Alignment for each Bureau ---
                for bureau in ["transunion", "experian", "equifax"]:
                    extracted_months = temp_extracted_months_per_bureau[bureau]
                    extracted_statuses = temp_extracted_statuses_per_bureau[bureau]

                    if extracted_statuses and "NONE REPORTED" in extracted_statuses:
                        account_data["payment_history"][bureau] = {"months": [], "statuses": "NONE REPORTED"}
                        print(f"✅ {bureau.title()} 2Y History for {creditor}: NONE REPORTED\n")
                        continue

                    max_history_length = 24

                    final_months = [''] * max_history_length
                    months_to_use = extracted_months[-max_history_length:]
                    
                    final_statuses = ['NO DATA'] * max_history_length
                    statuses_to_use = extracted_statuses[-max_history_length:]
                    
                    # Populate final_months and final_statuses from the right
                    for i in range(len(months_to_use)):
                        final_months[max_history_length - len(months_to_use) + i] = months_to_use[i].replace("'", "")
                    
                    for i in range(len(statuses_to_use)):
                        final_statuses[max_history_length - len(statuses_to_use) + i] = statuses_to_use[i]
                    
                    # This warning is for debugging purposes if you still see unexpected lengths
                    if len(months_to_use) != len(statuses_to_use):
                         print(f"⚠️ Warning: Mismatch in final aligned months ({len(months_to_use)}) and statuses ({len(statuses_to_use)}) for {bureau} - {creditor}. Padding 'NO DATA' for missing statuses.")

                    account_data["payment_history"][bureau] = {
                        "months": final_months,
                        "statuses": final_statuses
                    }
                    print(f"✅ {bureau.title()} 2Y History for {creditor}:\nMonths: {final_months}\nStatuses: {final_statuses}\n")

            else:
                print(f"❌ No 'Two-Year Payment History' section found for {creditor}")
        except Exception as e:
            print(f"⚠️ Error extracting payment history for {creditor}: {e}")
            import traceback
            traceback.print_exc()

        return account_data


    
    def extract_public_information(self, text):
        print("🔍 Extracting public information...")

        # Locate the public info section
        match = re.search(r'Public Information(.*?)(?:Collections|Inquiries|Accounts|End of Report)', text, re.DOTALL | re.IGNORECASE)
        if not match:
            print("❌ Public Information section not found.")
            return

        section = match.group(1)
        bureaus = ['transunion', 'experian', 'equifax']
        self.credit_report["public_information"] = {
            "transunion": {}, "experian": {}, "equifax": {}
        }

        # Special handling for 'Type' line (grab 9 words total -> 3 per bureau)
        type_line_match = re.search(r'Type\s+((?:\S+\s+){8}\S+)', section)
        if type_line_match:
            words = type_line_match.group(1).strip().split()
            if len(words) >= 9:
                self.credit_report["public_information"]["transunion"]["type"] = " ".join(words[0:3])
                self.credit_report["public_information"]["experian"]["type"] = " ".join(words[3:6])
                self.credit_report["public_information"]["equifax"]["type"] = " ".join(words[6:9])

                print(f"✅ Transunion - Type: {' '.join(words[0:3])}")
                print(f"✅ Experian - Type: {' '.join(words[3:6])}")
                print(f"✅ Equifax - Type: {' '.join(words[6:9])}")
            else:
                print("⚠️ Not enough words found for public info 'Type'")
        else:
            print("⚠️ Type line not found")

        # Process remaining fields using previous method
        fields = {
            'status': r'Status\s+(.*?)\s+(.*?)\s+(.*?)\n',
            'date_filed': r'Date Filed/Reported\s+(.*?)\s+(.*?)\s+(.*?)\n',
            'reference': r'Reference#\s+(.*?)\s+(.*?)\s+(.*?)\n'
        }

        for field, pattern in fields.items():
            m = re.search(pattern, section, re.IGNORECASE)
            if m:
                for i, bureau in enumerate(bureaus):
                    value = m.group(i + 1).strip()
                    self.credit_report["public_information"][bureau][field] = value
                    print(f"✅ {bureau.title()} - {field.replace('_', ' ').title()}: {value}")
            else:
                print(f"⚠️ Pattern not found for field: {field}")


    
    def clean_value(self, value):
        """Clean and standardize values"""
        if not value or value == '--' or value == 'N/A':
            return ''
        
        # Remove currency symbols and commas
        value = str(value).replace('$', '').replace(',', '')
        
        # Standardize date formats
        date_patterns = [
            (r'^\d{1,2}/\d{1,2}/\d{4}$', '%m/%d/%Y'),  # MM/DD/YYYY
            (r'^\d{1,2}/\d{1,2}/\d{2}$', '%m/%d/%y'),   # MM/DD/YY
            (r'^[A-Za-z]+\s+\d{1,2},\s*\d{4}$', '%B %d, %Y')  # Month DD, YYYY
        ]
        
        for pattern, date_format in date_patterns:
            if re.match(pattern, value):
                try:
                    date_obj = datetime.strptime(value, date_format)
                    return date_obj.strftime('%m/%d/%Y')  # Standardize to MM/DD/YYYY
                except ValueError:
                    pass
        
        return value
    
    def save_json(self):
        """Save the credit report data to a JSON file"""
        with open(self.json_output, 'w', encoding='utf-8') as f:
            json.dump(self.credit_report, f, indent=2)
        print(f"JSON data saved to {self.json_output}")
    
    def create_excel(self):
        """Create an Excel workbook with multiple sheets for the credit report"""
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Create all sheets
        self.create_summary_sheet(wb)
        self.create_inquiries_sheet(wb)
        
        for bureau in ["transunion", "experian", "equifax"]:
            self.create_bureau_sheet(wb, bureau)
        
        wb.save(self.excel_output)
        print(f"Excel file created at {self.excel_output}")

    def create_summary_sheet(self, workbook):
        """Create a summary sheet with credit scores and account counts"""
        ws = workbook.create_sheet("Summary")
        
        # Add title
        ws.merge_cells('A1:D1')
        ws['A1'] = "CREDIT REPORT SUMMARY"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add credit scores
        ws['A3'] = "Credit Scores"
        ws['A3'].font = Font(bold=True)
        
        ws['A4'] = "Bureau"
        ws['B4'] = "TransUnion"
        ws['C4'] = "Experian"
        ws['D4'] = "Equifax"
        
        ws['A5'] = "Score"
        ws['B5'] = self.credit_report["personal_info"]["transunion"].get("credit_score", "N/A")
        ws['C5'] = self.credit_report["personal_info"]["experian"].get("credit_score", "N/A")
        ws['D5'] = self.credit_report["personal_info"]["equifax"].get("credit_score", "N/A")
        
        # Add account summary
        ws['A7'] = "Account Summary"
        ws['A7'].font = Font(bold=True)
        
        # Summary rows to display
        summary_rows = [
            ("Total Accounts", "total_accounts"),
            ("Open Accounts", "open_accounts"),
            ("Closed Accounts", "closed_accounts"),
            ("Delinquent", "delinquent"),
            ("Derogatory", "derogatory"),
            ("Total Balances", "balances"),
            ("Monthly Payments", "monthly_payments"),
            ("Credit Utilization", "credit_utilization"),
            ("Public Records", "public_records"),
            ("Inquiries (2 years)", "inquiries_2y")
        ]
        
        row = 8
        for label, key in summary_rows:
            ws[f'A{row}'] = label
            
            for col, bureau in zip(['B', 'C', 'D'], ["transunion", "experian", "equifax"]):
                value = self.credit_report["summary"][bureau].get(key, "N/A")
                if key in ["balances", "monthly_payments"] and value != "N/A":
                    value = f"${value}"
                elif key == "credit_utilization" and value != "N/A":
                    value = f"{value}%"
                ws[f'{col}{row}'] = value
                
            row += 1
        
        # Set column widths
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20
    
    def create_inquiries_sheet(self, workbook):
        """Create a sheet for credit inquiries"""
        ws = workbook.create_sheet("Inquiries") 
        
        # Add title
        ws.merge_cells('A1:C1')
        ws['A1'] = "CREDIT INQUIRIES"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add headers
        headers = ["Creditor Name", "Date of Inquiry", "Credit Bureau"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")
        
        # Add data
        row_idx = 4
        for inquiry in self.credit_report["inquiries"]:
            ws.cell(row=row_idx, column=1, value=inquiry["creditor"])
            ws.cell(row=row_idx, column=2, value=inquiry["date"])
            ws.cell(row=row_idx, column=3, value=inquiry["bureau"].title())
            row_idx += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
    
    def create_bureau_sheet(self, workbook, bureau):
        """Create a sheet for a specific bureau's account data"""
        bureau_title = bureau.title()
        ws = workbook.create_sheet(bureau_title)
        
        # Define headers with all the fields
        headers = [
            "Creditor", 
            "Account Number", 
            "Account Type",
            "Balance Owed",
            "High Balance",
            "Credit Limit",
            "Account Status",
            "Payment Status",
            "Date Opened",
            "Last Payment",
            "Date Reported",
            "Term Length",
            "Past Due Amount",
            "Creditor Type",
            "Creditor Remarks",
            "Last Verified",
            "Date of Last Activity",
            "Closed Date",
            "Account Rating",
            "Account Description",
            "Dispute Status",
            "Payment Amount",
            "Payment Frequency",
            "Payment History"
        ]
        
        # Add title
        merged_cells = f'A1:{openpyxl.utils.get_column_letter(len(headers))}1'
        ws.merge_cells(merged_cells)
        ws['A1'] = f"{bureau_title} ACCOUNTS"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Add account data
        row_idx = 4
        for account in self.credit_report["accounts"]:
            if bureau in account and account[bureau]:
                # Set creditor name
                ws.cell(row=row_idx, column=1, value=account["creditor"])
                
                # Map the rest of the fields from bureau_data
                bureau_data = account[bureau]
                field_map = {
                    "account_number": 2,
                    "account_type": 3,
                    "balance_owed": 4,
                    "high_balance": 5,
                    "credit_limit": 6,
                    "account_status": 7,
                    "payment_status": 8,
                    "date_opened": 9,
                    "last_payment": 10,
                    "date_reported": 11,
                    "term_length": 12,
                    "past_due_amount": 13,
                    "creditor_type": 14,
                    "creditor_remarks": 15,
                    "last_verified": 16,
                    "date_of_last_activity": 17,
                    "closed_date": 18,
                    "account_rating": 19,
                    "account_description": 20,
                    "dispute_status": 21,
                    "payment_amount": 22,
                    "payment_frequency": 23,
                    "payment_history": 24
                }
                
                for field, col_idx in field_map.items():
                    value = bureau_data.get(field, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)
                
                row_idx += 1
        
        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_length = 0
            
            # Find the maximum content length in the column
            for row_idx in range(3, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            # Add a little extra space and cap at reasonable width
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[col_letter].width = adjusted_width


def main():
    # Default values
    pdf_path = "/Users/Alwinsaji/Downloads/newproject4/Credit report 0- Mylashia Monae Montgomery.pdf"
    output_dir = "output"

    # Use command-line arguments if provided
    if len(sys.argv) > 1:
        pdf_path = sys.argv[1]
    if len(sys.argv) > 2:
        output_dir = sys.argv[2]

    processor = CreditReportProcessor()
    processor.process(pdf_path, output_dir)

if __name__ == "__main__": 
    main()